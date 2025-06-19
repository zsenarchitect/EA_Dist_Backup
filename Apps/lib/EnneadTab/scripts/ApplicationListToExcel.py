import os
import sys
import subprocess
import importlib.util
import socket
import traceback

def check_and_install_module(module_name):
    """Check if a module is installed, if not install it."""
    if importlib.util.find_spec(module_name) is None:
        print(f"Module {module_name} not found. Installing...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", module_name])
            print(f"Successfully installed {module_name}")
        except subprocess.CalledProcessError as e:
            print(f"Error installing {module_name}: {e}")
            sys.exit(1)

# Check and install required modules
required_modules = [
    'pandas',
    'plotly',
    'openpyxl',
    'fuzzywuzzy'
]

for module in required_modules:
    check_and_install_module(module)

# Now import the modules after ensuring they are installed
import json
import glob
from pathlib import Path
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import webbrowser
from collections import Counter
import re
import openpyxl
import datetime
import textwrap
import time
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

# Priority application patterns that should appear first
PRIORITY_PATTERNS = [
    r"^Revit 2026$", r"^Autodesk Revit 2026$",
    r"^Revit 2025$", r"^Autodesk Revit 2025$",
    r"^Revit 2024$", r"^Autodesk Revit 2024$",
    r"^Revit 2023$", r"^Autodesk Revit 2023$",
    r"^Revit 2022$", r"^Autodesk Revit 2022$",
    r"^Revit 2021$", r"^Autodesk Revit 2021$",
    r"^Revit 2020$", r"^Autodesk Revit 2020$",
    r"^Revit 2019$", r"^Autodesk Revit 2019$",
    r"^Revit 2018$", r"^Autodesk Revit 2018$",
    r"^Revit 2017$", r"^Autodesk Revit 2017$",
    r"^Revit 2016$", r"^Autodesk Revit 2016$",
    r"^Revit 2015$", r"^Autodesk Revit 2015$",
    r"^Revit 2014$", r"^Autodesk Revit 2014$",
    r"^Rhino 9$", r"^Rhino 8$", r"^Rhino 7$", r"^Rhino 6$", r"^Rhino 5$", r"^Rhino 4$", r"^Rhino 3$", r"^Rhino 2$", r"^Rhino 1$"
]

def get_priority_score(app_name):
    """Calculate priority score for an application name.
    Higher score means higher priority."""
    for i, pattern in enumerate(PRIORITY_PATTERNS):
        if re.search(pattern, app_name, re.IGNORECASE):
            return i
    # Return a large number for non-priority apps
    return 999

def sort_applications(apps_dict):
    """Sort applications based on priority patterns and then alphabetically."""
    # Convert to list of tuples (app_name, priority_score)
    apps_with_scores = [(app_name, get_priority_score(app_name)) 
                       for app_name in apps_dict.keys()]
    
    # Sort first by priority score, then alphabetically
    sorted_apps = [app_name for app_name, _ in 
                  sorted(apps_with_scores, key=lambda x: (x[1], x[0].lower()))]
    
    return sorted_apps

def read_application_json_files():
    # Network share path where JSON files are stored
    share_path = r"L:\4b_Applied Computing\EnneadTab-DB\Shared Data Dump"
    
    # Check if path exists
    if not os.path.exists(share_path):
        print(f"Error: Path {share_path} does not exist!")
        return []
    
    # Get all JSON files matching the pattern
    json_files = glob.glob(os.path.join(share_path, "APPVERSIONLOOKUP_*.json"))
    
    if not json_files:
        print("No JSON files found!")
        return []
    
    print(f"Found {len(json_files)} JSON files:")
    # for file in json_files:
    #     print(f"- {os.path.basename(file)}")
    
    # Read and parse all JSON files
    all_data = []
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8-sig') as f:  # Using utf-8-sig to handle BOM
                data = json.load(f)
                all_data.append(data)
                # print(f"\nSuccessfully read {os.path.basename(json_file)}:")
        except Exception as e:
            print(f"Error reading {json_file}: {str(e)}")
    
    return all_data

def create_visualization(data, output_path):
    num_pcs = len(set(record['PC'] for record in data))
    generated_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Base HTML template with dedent
    base_html = textwrap.dedent(f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset=\"UTF-8\">
        <title>Application Usage Summary</title>
        <script src='https://cdn.plot.ly/plotly-latest.min.js'></script>
        <style>
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                margin: 0;
                background: #181a1b;
                color: #e0e0e0;
            }}
            .container {{
                max-width: 1400px;
                margin: 0 auto;
                background: #23272a;
                padding: 30px 30px 80px 30px;
                border-radius: 12px;
                box-shadow: 0 0 24px 0 #000a;
                min-height: 100vh;
            }}
            h1 {{
                color: #fff;
                text-align: center;
                margin-bottom: 30px;
            }}
            .tab {{
                overflow: hidden;
                border: 1px solid #333;
                background: #23272a;
                border-radius: 6px;
                margin-bottom: 24px;
            }}
            .tab button {{
                background: inherit;
                float: left;
                border: none;
                outline: none;
                cursor: pointer;
                padding: 16px 20px;
                transition: 0.3s;
                font-size: 17px;
                color: #bbb;
            }}
            .tab button:hover {{
                background: #333;
                color: #fff;
            }}
            .tab button.active {{
                background: #007bff;
                color: #fff;
            }}
            .tabcontent {{
                display: none;
                padding: 0;
                border: none;
                border-radius: 6px;
                background: transparent;
                animation: fadeIn 0.5s;
            }}
            .charts-grid {{
                display: flex;
                flex-wrap: wrap;
                gap: 24px;
                justify-content: flex-start;
            }}
            .chart-card {{
                background: #23272a;
                border-radius: 10px;
                box-shadow: 0 2px 12px #0006;
                padding: 18px 12px 12px 12px;
                margin-bottom: 0;
                flex: 1 1 calc(33.333% - 24px);
                min-width: 350px;
                max-width: 420px;
                transition: box-shadow 0.2s, transform 0.2s;
                position: relative;
            }}
            .chart-card:hover {{
                box-shadow: 0 6px 24px #007bff99;
                transform: translateY(-4px) scale(1.02);
            }}
            .chart-title {{
                color: #fff;
                font-size: 1.1em;
                margin-bottom: 10px;
                text-align: center;
                font-weight: 500;
                letter-spacing: 0.5px;
                transition: background 0.3s;
                padding: 2px 0;
            }}
            .highlight {{
                background: #2a2d3a;
                color: #ffd700;
                border-radius: 4px;
            }}
            .back-to-top {{
                position: fixed;
                bottom: 30px;
                right: 30px;
                background: #007bff;
                color: #fff;
                border: none;
                border-radius: 50%;
                width: 48px;
                height: 48px;
                font-size: 24px;
                cursor: pointer;
                box-shadow: 0 2px 8px #0008;
                z-index: 1000;
                display: none;
                transition: background 0.3s, box-shadow 0.3s;
            }}
            .back-to-top:hover {{
                background: #0056b3;
                box-shadow: 0 4px 16px #007bff99;
            }}
            @keyframes fadeIn {{
                from {{ opacity: 0; }}
                to {{ opacity: 1; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Application Usage Summary ({num_pcs} PCs)</h1>
            <p style='text-align:center; color:#bbb; font-size:0.95em; margin-top:-18px; margin-bottom:10px;'>
                Report generated: {generated_time}
            </p>
            <p style='text-align:center; color:#bbb; font-size:1em; font-style:italic; margin-top:-10px; margin-bottom:30px;'>
              All applications with similar names will be displayed separately here. For example, "Revit 2024" and "Autodesk Revit 2024" are recorded separately by the computer system.
            </p>
            <div class="tab">
                <button class="tablinks active" onclick="openTab(event, 'Revit')">Revit</button>
                <button class="tablinks" onclick="openTab(event, 'Rhino')">Rhino</button>
                <button class="tablinks" onclick="openTab(event, 'Enscape')">Enscape</button>
            </div>
    """)
    
    # Initialize html_content with base template
    html_content = base_html
    
    # Process data for each application type
    for app_type in ['Revit', 'Rhino', 'Enscape']:
        html_content += f'<div id="{app_type}" class="tabcontent" style="display: none;">'
        html_content += '<div class="charts-grid">'
        version_data = {}
        for record in data:
            if app_type in record:
                for app_name, version in record[app_type].items():
                    if app_name not in version_data:
                        version_data[app_name] = Counter()
                    version_data[app_name][version] += 1
        sorted_apps = sort_applications(version_data)
        # Debug output: print app names and their priority scores
        # print(f"\n[{app_type}] Sorted app order and priority scores:")
        for app_name in sorted_apps:
            score = get_priority_score(app_name)
            # print(f"  {app_name} (score: {score})")
        for app_name in sorted_apps:
            versions = version_data[app_name]
            chart_id = f"{app_type}_{app_name.replace(' ', '_').replace('.', '_').replace('(', '').replace(')', '')}"
            html_content += f'<div class="chart-card"><div class="chart-title" id="title_{chart_id}">{app_name}</div>'
            html_content += f'<div id="{chart_id}" class="chart-container"></div></div>'
            labels = list(versions.keys())
            values = list(versions.values())
            html_content += f"""
            <script>
                var data = [{{
                    values: {values},
                    labels: {labels},
                    type: 'pie',
                    name: '{app_name}',
                    hoverinfo: 'label+value+percent',
                    textinfo: 'label+value',
                    textposition: 'outside',
                    automargin: true,
                    marker: {{
                        colors: ['#007bff', '#17a2b8', '#ffc107', '#dc3545', '#6610f2', '#fd7e14', '#20c997', '#6f42c1', '#343a40', '#adb5bd']
                    }},
                    textfont: {{
                        size: 14,
                        color: '#e0e0e0'
                    }},
                    insidetextorientation: 'radial'
                }}];
                var layout = {{
                    showlegend: true,
                    legend: {{
                        orientation: 'h',
                        y: -0.1,
                        font: {{color: '#e0e0e0'}}
                    }},
                    height: 340,
                    paper_bgcolor: '#23272a',
                    plot_bgcolor: '#23272a',
                    font: {{color: '#e0e0e0'}},
                    margin: {{t: 30, b: 30, l: 30, r: 30}}
                }};
                var config = {{
                    displayModeBar: true,
                    displaylogo: false,
                    responsive: true,
                    animation: {{
                        duration: 1000,
                        easing: 'cubic-in-out'
                    }}
                }};
                Plotly.newPlot('{chart_id}', data, layout, config);
            </script>
            """
        html_content += '</div></div>'
    
    # Add closing HTML and JavaScript
    html_content += textwrap.dedent("""
        <button class='back-to-top' id='backToTopBtn' title='Back to top'>&uarr;</button>
        <script>
        function openTab(evt, appType) {
            var i, tabcontent, tablinks;
            tabcontent = document.getElementsByClassName('tabcontent');
            for (i = 0; i < tabcontent.length; i++) {
                tabcontent[i].style.display = 'none';
            }
            tablinks = document.getElementsByClassName('tablinks');
            for (i = 0; i < tablinks.length; i++) {
                tablinks[i].className = tablinks[i].className.replace(' active', '');
            }
            var tab = document.getElementById(appType);
            if (tab) tab.style.display = 'block';
            if (evt) {
                evt.currentTarget.className += ' active';
            } else {
                for (i = 0; i < tablinks.length; i++) {
                    if (tablinks[i].textContent === appType) {
                        tablinks[i].className += ' active';
                        break;
                    }
                }
            }
            window.dispatchEvent(new Event('resize'));
        }
        document.addEventListener('DOMContentLoaded', function() {
            openTab(null, 'Revit');
        });
        var backToTopBtn = document.getElementById('backToTopBtn');
        window.onscroll = function() {
            if (document.body.scrollTop > 200 || document.documentElement.scrollTop > 200) {
                backToTopBtn.style.display = 'block';
            } else {
                backToTopBtn.style.display = 'none';
            }
        };
        backToTopBtn.onclick = function() {
            window.scrollTo({top: 0, behavior: 'smooth'});
        };
        </script>
    </body>
    </html>
    """)
    
    html_path = os.path.join(output_path, "ApplicationUsageSummary.html")
    try:
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    except PermissionError:
        print(f"[ERROR] Permission denied: Could not write to {html_path}. Please close the file if it is open and try again.")
        return None
    except Exception as e:
        print(f"[ERROR] Could not write HTML file: {e}")
        return None
    return html_path

def create_excel_report(data, output_path, matches=None, unmatched_employees=None):
    print("[LOG] Starting Excel report creation...")
    # Group records by PC
    pc_records = {}
    pc_user_pairs = []
    pcs_with_data = set()
    for record in data:
        pc = record['PC']
        user = record['User']
        if pc not in pc_records:
            pc_records[pc] = []
            pc_user_pairs.append((pc, user))
        for app_type in ['Revit', 'Rhino', 'Enscape']:
            if app_type in record:
                for app_name, version in record[app_type].items():
                    pc_records[pc].append({
                        'Application Type': app_type,
                        'Application Name': app_name,
                        'Version': version,
                        'User': user
                    })
                    pcs_with_data.add(pc)

    # Prepare comment column
    pc_user_comments = []
    for pc, user in pc_user_pairs:
        comment = "" if pc in pcs_with_data else "No application data found"
        pc_user_comments.append((pc, user, comment))

    # Generate timestamp for filename
    timestamp = datetime.datetime.now().strftime('%Y%m%d-%H%M')
    excel_path = os.path.join(output_path, f"ApplicationUsageSummary_{timestamp}.xlsx")
    print(f"[LOG] Excel path: {excel_path}")
    try:
        print("[LOG] Writing Excel file with pandas...")
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_pc_users = pd.DataFrame(pc_user_comments, columns=["ComputerName", "UserName", "Comment"])
            df_pc_users.to_excel(writer, sheet_name="PC UserNames", index=False)

            # Insert Employees With Data (2nd sheet)
            if matches is not None:
                df_with_data = pd.DataFrame([
                    {'Employee Name': m['employee_name'], 'Profile Username': m['username'], 'How confident i am feeling my guess is correct': m['confidence']} for m in matches
                ])
                df_with_data = df_with_data.sort_values(by='How confident i am feeling my guess is correct', ascending=False)
                df_with_data.to_excel(writer, sheet_name="Employees With Data", index=False)

            # Insert Employees Without Data (3rd sheet)
            if unmatched_employees is not None:
                df_without_data = pd.DataFrame({'Employee Name': unmatched_employees})
                df_without_data.to_excel(writer, sheet_name="Employees Without Data", index=False)

            for pc, records in pc_records.items():
                df = pd.DataFrame(records)
                if not df.empty:
                    df = df.sort_values(['Application Type', 'Application Name'], key=lambda col: col.str.lower())
                    safe_pc = str(pc)[:31]
                    df.to_excel(writer, sheet_name=safe_pc, index=False)
                else:
                    print(f"[LOG] Skipping empty sheet for PC: {pc}")
        print("[LOG] Excel file written, now formatting with openpyxl...")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["PC UserNames"]
        wb.active = wb.sheetnames.index("PC UserNames")
        from openpyxl.styles import Font, Border, Side, PatternFill
        from openpyxl.formatting.rule import FormulaRule
        print("[LOG] Formatting header...")
        for cell in ws[1]:
            cell.font = Font(bold=True)
        print("[LOG] Adding borders...")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
        print("[LOG] Adding conditional formatting...")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        max_row = ws.max_row
        formula = 'ISNUMBER(SEARCH("EANY",A2))'
        ws.conditional_formatting.add(f"A2:A{max_row}", FormulaRule(formula=[formula], fill=green_fill))
        print(f"[LOG] Applied conditional formatting to range A2:A{max_row} with formula: {formula}")
        ws.auto_filter.ref = "A1:C1"
        print(f"[LOG] Applied autofilter to range A1:C1")
        print("[LOG] Auto-sizing columns...")
        for ws_sheet in wb.worksheets:
            for col in ws_sheet.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_sheet.column_dimensions[col_letter].width = adjusted_width
        print("[LOG] Saving workbook...")
        wb.save(excel_path)
        print("[LOG] Excel report successfully saved.")
    except PermissionError:
        print(f"[ERROR] Permission denied: Could not write to {excel_path}. Please close the file if it is open and try again.")
        return None
    except Exception as e:
        print(f"[ERROR] Could not write Excel file: {e}")
        traceback.print_exc()
        return None
    return excel_path

def read_employee_names():
    """Read employee names and email usernames from the NY Employees Excel file."""
    employee_file = r"L:\4b_Applied Computing\EnneadTab-DB\Shared Data Dump\NY Employees.xlsx"
    
    if not os.path.exists(employee_file):
        print(f"Error: Employee file {employee_file} does not exist!")
        return []
    
    try:
        df = pd.read_excel(employee_file)
        if df.empty:
            print("Error: Employee file is empty!")
            return []
        # Expecting column A: email, column B: full name
        email_col = df.columns[0]
        name_col = df.columns[1] if len(df.columns) > 1 else None
        employees = []
        for idx, row in df.iterrows():
            email = str(row[email_col]).strip()
            full_name = str(row[name_col]).strip() if name_col else ""
            if not email or email.lower() == 'nan':
                continue
            email_username = email.split('@')[0]
            employees.append({'full_name': full_name, 'email_username': email_username})
        print(f"Found {len(employees)} employee records (email + name)")
        return employees
    except Exception as e:
        print(f"Error reading employee file: {e}")
        return []

def normalize_name(name):
    """Normalize a name for better matching."""
    if not name:
        return ""
    # Convert to lowercase and remove extra spaces
    normalized = str(name).lower().strip()
    # Remove common prefixes/suffixes that might appear in usernames
    normalized = re.sub(r'^[a-z]\.', '', normalized)  # Remove single letter prefix like "j."
    normalized = re.sub(r'[^\w\s]', '', normalized)   # Remove special characters except spaces
    return normalized

def extract_last_name(full_name):
    """Extract last name from full name."""
    if not full_name:
        return ""
    parts = str(full_name).strip().split()
    return parts[-1].lower() if parts else ""

def compare_names_with_usernames(employee_records, pc_usernames, threshold=90):
    """Compare employee email usernames and names with PC usernames using exact and fuzzy matching. Only matches with confidence >= 90 are considered valid."""
    matches = []
    unmatched_employees = []
    unmatched_usernames = list(pc_usernames)
    
    print(f"\nComparing {len(employee_records)} employee records with {len(pc_usernames)} PC usernames...")
    print(f"Using threshold: {threshold}%")
    
    for emp in employee_records:
        employee_name = emp['full_name']
        email_username = emp['email_username']
        best_match = None
        best_score = 0
        # 1. Try exact match with email username
        for username in pc_usernames:
            if username.lower() == email_username.lower():
                best_match = username
                best_score = 100
                break
        # 2. If no exact match, use fuzzy matching on full name
        if best_score < 100:
            norm_employee = normalize_name(employee_name)
            employee_last_name = extract_last_name(employee_name)
            for username in pc_usernames:
                norm_username = normalize_name(username)
                scores = []
                scores.append(fuzz.ratio(norm_employee, norm_username))
                scores.append(fuzz.partial_ratio(norm_employee, norm_username))
                scores.append(fuzz.token_sort_ratio(norm_employee, norm_username))
                if employee_last_name and employee_last_name in norm_username:
                    scores.append(85)
                if len(employee_name.split()) >= 2:
                    first_initial = employee_name.split()[0][0].lower()
                    if norm_username.startswith(first_initial) and employee_last_name in norm_username:
                        scores.append(90)
                max_score = max(scores)
                if max_score > best_score:
                    best_score = max_score
                    best_match = username
        if best_score >= threshold:
            matches.append({
                'employee_name': employee_name,
                'username': best_match,
                'confidence': best_score
            })
            if best_match in unmatched_usernames:
                unmatched_usernames.remove(best_match)
            print(f"✓ MATCH: '{employee_name}' (email: {email_username}) -> '{best_match}' (confidence: {best_score}%)")
        else:
            unmatched_employees.append(employee_name)
            print(f"✗ NO MATCH: '{employee_name}' (email: {email_username}) (best: '{best_match}' at {best_score}%)")
    print(f"\n=== MATCHING SUMMARY ===")
    print(f"Total matches found: {len(matches)}")
    print(f"Unmatched employees: {len(unmatched_employees)}")
    print(f"Unmatched usernames: {len(unmatched_usernames)}")
    if unmatched_employees:
        print(f"\nUnmatched employees:")
        for name in unmatched_employees:
            print(f"  - {name}")
    if unmatched_usernames:
        print(f"\nUnmatched usernames:")
        for username in unmatched_usernames:
            print(f"  - {username}")
    return matches, unmatched_employees, unmatched_usernames

def analyze_missing_data(data, employee_matches):
    """Analyze which employees might be missing data in PC UserNames."""
    # Get all PC usernames from the data
    pc_usernames = set()
    for record in data:
        pc_usernames.add(record['User'])
    
    # Get employee names that have matches
    matched_employees = {match['employee_name'] for match in employee_matches}
    
    # Find employees who might be missing data
    missing_data_employees = []
    for match in employee_matches:
        employee_name = match['employee_name']
        username = match['username']
        
        # Check if this username appears in the data
        user_found = False
        for record in data:
            if record['User'] == username:
                user_found = True
                break
        
        if not user_found:
            missing_data_employees.append({
                'employee_name': employee_name,
                'username': username,
                'confidence': match['confidence']
            })
    
    print(f"\n=== MISSING DATA ANALYSIS ===")
    print(f"Employees potentially missing data: {len(missing_data_employees)}")
    
    for item in missing_data_employees:
        print(f"  - {item['employee_name']} (username: {item['username']}, confidence: {item['confidence']}%)")
    
    return missing_data_employees

def main(open_after = True):
    print("Reading application JSON files...")
    data = read_application_json_files()
    print(f"\nTotal records processed: {len(data)}")
    
    matches = None
    unmatched_employees = None
    if data:
        output_path = r"L:\4b_Applied Computing\EnneadTab-DB\Shared Data Dump"
        
        # Employee name comparison
        print("\n" + "="*60)
        print("EMPLOYEE NAME COMPARISON")
        print("="*60)
        
        # Read employee names
        employee_names = read_employee_names()
        if employee_names:
            # Get PC usernames from the data
            pc_usernames = set()
            for record in data:
                pc_usernames.add(record['User'])
            
            # Compare names
            matches, unmatched_employees, unmatched_usernames = compare_names_with_usernames(
                employee_names, pc_usernames, threshold=90
            )
            
            # Analyze missing data
            missing_data_employees = analyze_missing_data(data, matches)
            
            # Print final summary of matches
            print(f"\n" + "="*60)
            print("FINAL MATCHES SUMMARY")
            print("="*60)
            print("Profile Username -> Full Human Name pairs:")
            for match in matches:
                print(f"  '{match['username']}' -> '{match['employee_name']}' (confidence: {match['confidence']}%)")
        
        # Create HTML visualization
        html_path = create_visualization(data, output_path)
        if html_path:
            print(f"\nHTML report created: {html_path}")
        
        # Create Excel report (with new sheets if matches info is available)
        excel_path = create_excel_report(data, output_path, matches=matches, unmatched_employees=unmatched_employees)
        if excel_path:
            print(f"Excel report created: {excel_path}")

        # Open HTML in default browser
        if open_after:
            webbrowser.open('file://' + html_path)




if __name__ == "__main__":
    computer_name = socket.gethostname().upper()
    if computer_name == "EANY-VIRTUAL7-1":
        print(f"Running in loop mode on {computer_name}.")
        while True:
            main(open_after = False)
            print("Sleeping for 30 minutes before next run...")
            for remaining in range(1800, 0, -1):
                mins, secs = divmod(remaining, 60)
                timer = '{:02d}:{:02d}'.format(mins, secs)
                print(f'\033[33m\rNext run in: {timer}\033[0m', end='')
                time.sleep(1)
            print('\rNext run starting...     ')
    else:
        main(open_after = True)  
